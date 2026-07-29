"""Safe, idempotent import for Controle de Compras R00.xlsm.

Historical receipts are recorded without stock movement. The present balance is
never recalculated or altered by legacy imports.
"""
import argparse, json
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from sqlalchemy import text
from database import engine

EMPTY={'','-','0','N/A','NA','AG','?'}
def clean(v):
    if v is None:return ''
    x=str(v).strip(); return '' if x.upper() in EMPTY else x
def num(v):
    try:return Decimal(clean(v).replace('.','').replace(',','.')) if isinstance(v,str) and ',' in v else Decimal(str(v or 0))
    except (InvalidOperation,ValueError): return Decimal('0')
def date(v): return v if isinstance(v,datetime) else None
def headers(ws): return {clean(ws.cell(2,c).value).upper().replace('\n',' '):c-1 for c in range(1,ws.max_column+1) if clean(ws.cell(2,c).value)}
def col(values,h,*names):
    for n in names:
        i=h.get(n.upper())
        if i is not None:return values[i]
    return None
def key(file,sheet,row):return f'{Path(file).name}:{sheet}:{row}'
def already(conn,k):return conn.execute(text('select 1 from erp_legacy_import_records where source_key=:k'),{'k':k}).first()
def record(conn,k,file,sheet,row,kind,eid,payload):
    conn.execute(text("insert into erp_legacy_import_records(source_key,source_file,source_sheet,source_item,entity_type,entity_id,payload) values(:k,:f,:s,:r,:kind,:id,cast(:p as jsonb))"),{'k':k,'f':Path(file).name,'s':sheet,'r':str(row),'kind':kind,'id':eid,'p':json.dumps(payload,ensure_ascii=False,default=str)})

def import_orders(conn,file,sheet,category,dry,report):
    ws=load_workbook(file,read_only=True,data_only=True,keep_links=False)[sheet]; h=headers(ws)
    for row,values in enumerate(ws.iter_rows(min_row=3,values_only=True),3):
        order=clean(col(values,h,'Nº PEDIDO ')); desc=clean(col(values,h,'DESCRIÇÃO','DESCRIÇÃO DO CÓDIGO'))
        if not order and not desc:continue
        k=key(file,sheet,row)
        if already(conn,k):report['ignored']+=1;continue
        qty=num(col(values,h,'QTD.','QTD')); value=num(col(values,h,'VALOR'))
        if qty<=0: report['rejected'].append({'sheet':sheet,'row':row,'error':'quantidade invalida'});continue
        report['orders']+=1
        if dry:continue
        po=str(uuid4()); line=str(uuid4()); supplier=clean(col(values,h,'FORNECEDOR'))
        status='CANCELADA' if clean(col(values,h,'STATUS')).upper()=='CANCELADO' else 'EMITIDA'
        conn.execute(text("insert into erp_purchase_orders(id,numero_oc,categoria,fornecedor_nome,data_criacao,data_emissao,criado_por,status,destino,frete,data_necessidade,observacoes,valor_total_pedido,idempotency_key) values(:id,:number,:category,:supplier,:created,:created,'IMPORTADOR',:status,:destino,0,:need,:obs,:value,:key)"),{'id':po,'number':order or f'LEG-{row}','category':category,'supplier':supplier,'created':date(col(values,h,'DATA PC')) or datetime.utcnow(),'status':status,'destino':clean(col(values,h,'DESTINO')),'need':date(col(values,h,'DATA NECESSIDADE ENTREGA')),'obs':clean(col(values,h,'OBSERVAÇÃO')),'value':value,'key':k})
        conn.execute(text("insert into erp_purchase_order_lines(id,purchase_order_id,numero_linha,descricao_original,unidade,quantidade_pedida,valor_unitario_pedido,destino,cliente_id,data_necessidade,status) values(:id,:po,1,:desc,'UN',:qty,:unit,:destino,:client,:need,:status)"),{'id':line,'po':po,'desc':desc or 'SEM DESCRICAO','qty':qty,'unit':value/qty,'destino':clean(col(values,h,'DESTINO')),'client':clean(col(values,h,'CLIENTE')) or None,'need':date(col(values,h,'DATA NECESSIDADE ENTREGA')),'status':'CANCELADA' if status=='CANCELADA' else 'PENDENTE'})
        record(conn,k,file,sheet,row,'PURCHASE_ORDER',po,{'numero_oc':order,'categoria':category})

def import_receipts(conn,file,dry,report):
    sheet='04 - Inspeção de Recebimento'; ws=load_workbook(file,read_only=True,data_only=True,keep_links=False)[sheet]; h=headers(ws)
    for row,values in enumerate(ws.iter_rows(min_row=3,values_only=True),3):
        supplier=clean(col(values,h,'FORNECEDOR')); description=clean(col(values,h,'DESCRIÇÃO'))
        if not supplier and not description:continue
        k=key(file,sheet,row)
        if already(conn,k):report['ignored']+=1;continue
        quantity=num(col(values,h,'QUANT. ENTREGUE'))
        if quantity<=0:report['rejected'].append({'sheet':sheet,'row':row,'error':'quantidade invalida'});continue
        report['receipts']+=1
        if dry:continue
        result=clean(col(values,h,'A','AC','D')).upper(); result=result if result in {'A','AC','D'} else 'A'
        receipt=str(uuid4()); line=str(uuid4()); approved=quantity if result=='A' else Decimal('0'); conditional=quantity if result=='AC' else Decimal('0'); rejected=quantity if result=='D' else Decimal('0')
        conn.execute(text("insert into erp_goods_receipts(id,origem,data_recebimento,fornecedor_nome,numero_nf,operador,status,observacoes,motivo_excecao,idempotency_key) values(:id,'MANUAL',:date,:supplier,:nf,:actor,'CONFIRMADO',:obs,'IMPORTADO COMO HISTORICO; SEM MOVIMENTO DE SALDO',:key)"),{'id':receipt,'date':date(col(values,h,'DATA ENTREGUE')) or datetime.utcnow(),'supplier':supplier,'nf':clean(col(values,h,'N° NF')),'actor':clean(col(values,h,'RESPONSÁVEL')) or 'IMPORTADOR','obs':clean(col(values,h,'OBSERVAÇÕES')),'key':k})
        conn.execute(text("insert into erp_goods_receipt_lines(id,goods_receipt_id,sku_codigo,quantidade_fisica,quantidade_aprovada,quantidade_condicional,quantidade_rejeitada,valor_unitario_real,resultado_inspecao,justificativa_divergencia) values(:id,:receipt,null,:qty,:approved,:conditional,:rejected,:value,:result,'LEGADO: sem movimento retroativo')"),{'id':line,'receipt':receipt,'qty':quantity,'approved':approved,'conditional':conditional,'rejected':rejected,'value':num(col(values,h,'VALOR TOTAL'))/quantity,'result':result})
        record(conn,k,file,sheet,row,'GOODS_RECEIPT',receipt,{'referencia':clean(col(values,h,'N° PC Invoice/ proforma','N° PC INVOICE/ PROFORMA')),'descricao':description})

def main():
    p=argparse.ArgumentParser();p.add_argument('file');p.add_argument('--dry-run',action='store_true');p.add_argument('--report',default='compras_import_report.json');a=p.parse_args();r={'dry_run':a.dry_run,'orders':0,'receipts':0,'ignored':0,'rejected':[]}
    with engine.begin() as c:
        import_orders(c,a.file,'01 - Controle Geral de Compras','GERAL',a.dry_run,r)
        import_orders(c,a.file,'02 - Controle de Compras Bancos','BANCOS',a.dry_run,r)
        import_receipts(c,a.file,a.dry_run,r)
    Path(a.report).write_text(json.dumps(r,ensure_ascii=False,indent=2,default=str),encoding='utf8');print(json.dumps(r,ensure_ascii=False,default=str))
if __name__=='__main__':main()

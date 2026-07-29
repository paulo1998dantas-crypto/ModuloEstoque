"""Relatório operacional consolidado de Compras e Inspeção."""
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text


PURCHASE_HEADERS = [
    "STATUS", "TIPO", "DATA PC", "Nº PEDIDO", "FORNECEDOR", "VALOR",
    "QTD.", "CÓDIGO", "DESCRIÇÃO", "DESTINO", "CLIENTE", "FRETE",
    "DATA NECESSIDADE ENTREGA", "DATA RECEBIMENTO",
    "TEMPO TOTAL (DATA P.C. x RECEBIMENTO)", "DIF. NECESS VS ENTREGA",
    "OBSERVAÇÃO",
]
INSPECTION_HEADERS = [
    "Data entregue", "Fornecedor", "N° PC / Invoice / proforma",
    "Descrição", "Quant. entregue", "N° NF", "Valor Total",
    "Responsável", "A", "AC", "D", "Observações",
]


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _days(start, end):
    start_date, end_date = _as_date(start), _as_date(end)
    return (end_date - start_date).days if start_date and end_date else None


def _purchase_status(order_status, ordered_quantity, received_quantity):
    if str(order_status or "") == "CANCELADA":
        return "CANCELADA"
    ordered = float(ordered_quantity or 0)
    received = float(received_quantity or 0)
    if received <= 0:
        return "AGUARDANDO"
    if received < ordered:
        return "ENTREGUE PARCIAL"
    return "ENTREGUE"


def _rows(db):
    purchases = [
        dict(row._mapping) for row in db.execute(text("""
            select o.numero_oc,o.categoria,o.status as order_status,
                   o.fornecedor_nome,o.data_emissao,o.destino,o.frete,
                   o.data_necessidade,o.observacoes,
                   l.numero_linha,l.sku_codigo,l.descricao_original,l.cliente_id,
                   l.quantidade_pedida,l.valor_unitario_pedido,
                   sum(l.quantidade_pedida) over (partition by o.id) quantidade_pedido,
                   sum(l.quantidade_recebida) over (partition by o.id) quantidade_recebida,
                   coalesce(l.data_necessidade,o.data_necessidade) necessidade_linha,
                   delivered.data_entregue
              from erp_purchase_orders o
              join erp_purchase_order_lines l on l.purchase_order_id=o.id
              left join lateral (
                  select max(r.data_recebimento) data_entregue
                    from erp_goods_receipt_lines rl
                    join erp_goods_receipts r on r.id=rl.goods_receipt_id
                    join erp_purchase_order_lines receipt_line
                      on receipt_line.id=rl.purchase_order_line_id
                   where receipt_line.purchase_order_id=o.id
                     and r.status='CONFIRMADO'
                     and coalesce(rl.quantidade_aprovada,0)
                       + coalesce(rl.quantidade_condicional,0) > 0
              ) delivered on true
             where coalesce(o.criado_por,'') <> 'validacao-local'
             order by o.data_emissao,o.numero_oc,l.numero_linha
        """)).all()
    ]
    inspections = [
        dict(row._mapping) for row in db.execute(text("""
            select r.data_recebimento,r.fornecedor_nome,r.numero_nf,r.operador,
                   r.observacoes receipt_notes,o.numero_oc,
                   coalesce(l.descricao_original,rl.sku_codigo,'ITEM SEM DESCRIÇÃO') descricao,
                   rl.quantidade_fisica,rl.valor_unitario_real,
                   l.valor_unitario_pedido,rl.resultado_inspecao,
                   rl.justificativa_divergencia
              from erp_goods_receipts r
              join erp_goods_receipt_lines rl on rl.goods_receipt_id=r.id
              left join erp_purchase_order_lines l on l.id=rl.purchase_order_line_id
              left join erp_purchase_orders o on o.id=r.purchase_order_id
             where r.status='CONFIRMADO'
               and coalesce(r.operador,'') <> 'validacao-local'
             order by r.data_recebimento,r.created_at,rl.id
        """)).all()
    ]
    return purchases, inspections


def _style_sheet(sheet, headers, row_count, title):
    navy = PatternFill("solid", fgColor="123D6A")
    teal = PatternFill("solid", fgColor="087F7A")
    thin = Side(style="thin", color="DCE4ED")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = navy
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    for cell in sheet[3]:
        cell.fill = teal
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[3].height = 38
    for row in sheet.iter_rows(min_row=4, max_row=row_count + 3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(row_count + 3, 4)}"
    widths = {
        "A": 20, "B": 12, "C": 13, "D": 17, "E": 25, "F": 15,
        "G": 12, "H": 16, "I": 42, "J": 22, "K": 25, "L": 16,
        "M": 18, "N": 18, "O": 24, "P": 22, "Q": 48,
    }
    for letter, width in widths.items():
        if sheet.max_column >= ord(letter) - 64:
            sheet.column_dimensions[letter].width = width


def build_purchase_inspection_report(db):
    purchases, inspections = _rows(db)
    workbook = Workbook()
    purchase_sheet = workbook.active
    purchase_sheet.title = "CONTROLE DE COMPRAS"
    for column, header in enumerate(PURCHASE_HEADERS, start=1):
        purchase_sheet.cell(3, column, header)
    for row_number, row in enumerate(purchases, start=4):
        delivered = _as_date(row.get("data_entregue"))
        need = _as_date(row.get("necessidade_linha"))
        values = [
            _purchase_status(
                row.get("order_status"),
                row.get("quantidade_pedido"),
                row.get("quantidade_recebida"),
            ),
            row.get("categoria") or "GERAL",
            _as_date(row.get("data_emissao")),
            row.get("numero_oc") or "",
            row.get("fornecedor_nome") or "",
            float(row.get("quantidade_pedida") or 0)
            * float(row.get("valor_unitario_pedido") or 0),
            float(row.get("quantidade_pedida") or 0),
            row.get("sku_codigo") or "",
            row.get("descricao_original") or "",
            row.get("destino") or "",
            row.get("cliente_id") or "",
            row.get("frete") or "",
            need,
            delivered,
            _days(row.get("data_emissao"), delivered),
            _days(need, delivered),
            row.get("observacoes") or "",
        ]
        for column, value in enumerate(values, start=1):
            purchase_sheet.cell(row_number, column, value)
    _style_sheet(
        purchase_sheet,
        PURCHASE_HEADERS,
        len(purchases),
        "CONTROLE DE COMPRAS — GERAL E BANCOS",
    )
    for column in ("C", "M", "N"):
        for cell in purchase_sheet[column][3:]:
            cell.number_format = "dd/mm/yyyy"
    for cell in purchase_sheet["F"][3:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in purchase_sheet["G"][3:]:
        cell.number_format = "#,##0.000"

    inspection_sheet = workbook.create_sheet("INSPEÇÃO DE RECEBIMENTO")
    for column, header in enumerate(INSPECTION_HEADERS, start=1):
        inspection_sheet.cell(3, column, header)
    for row_number, row in enumerate(inspections, start=4):
        result = str(row.get("resultado_inspecao") or "")
        notes = " | ".join(
            value for value in (
                str(row.get("receipt_notes") or "").strip(),
                str(row.get("justificativa_divergencia") or "").strip(),
            ) if value
        )
        values = [
            _as_date(row.get("data_recebimento")),
            row.get("fornecedor_nome") or "",
            row.get("numero_oc") or "ENTRADA MANUAL",
            row.get("descricao") or "",
            float(row.get("quantidade_fisica") or 0),
            row.get("numero_nf") or "",
            float(row.get("quantidade_fisica") or 0)
            * float(
                row.get("valor_unitario_real")
                or row.get("valor_unitario_pedido")
                or 0
            ),
            row.get("operador") or "",
            "A" if result == "A" else "",
            "AC" if result == "AC" else "",
            "D" if result == "D" else "",
            notes,
        ]
        for column, value in enumerate(values, start=1):
            inspection_sheet.cell(row_number, column, value)
    _style_sheet(
        inspection_sheet,
        INSPECTION_HEADERS,
        len(inspections),
        "INSPEÇÃO DE RECEBIMENTO",
    )
    inspection_sheet["A2"] = "Legenda:"
    inspection_sheet["B2"] = "A - Aprovado"
    inspection_sheet["C2"] = "AC - Aprovado Condicional"
    inspection_sheet["D2"] = "D - Devolver"
    inspection_sheet["A2"].font = Font(bold=True)
    for column in ("A",):
        for cell in inspection_sheet[column][3:]:
            cell.number_format = "dd/mm/yyyy"
    for cell in inspection_sheet["E"][3:]:
        cell.number_format = "#,##0.000"
    for cell in inspection_sheet["G"][3:]:
        cell.number_format = 'R$ #,##0.00'
    inspection_sheet.column_dimensions["A"].width = 16
    inspection_sheet.column_dimensions["B"].width = 28
    inspection_sheet.column_dimensions["C"].width = 24
    inspection_sheet.column_dimensions["D"].width = 42
    inspection_sheet.column_dimensions["E"].width = 16
    inspection_sheet.column_dimensions["F"].width = 18
    inspection_sheet.column_dimensions["G"].width = 16
    inspection_sheet.column_dimensions["H"].width = 20
    inspection_sheet.column_dimensions["I"].width = 8
    inspection_sheet.column_dimensions["J"].width = 8
    inspection_sheet.column_dimensions["K"].width = 8
    inspection_sheet.column_dimensions["L"].width = 48

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, len(purchases), len(inspections)

from datetime import datetime
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_

from config import EXPORTS_DIR
from models import BomComponent, InventoryCount, InventorySession, LabelPrintJob, Movement, SKU, StockBalance
from services.etiqueta_service import create_label_job
from services.estoque_service import (
    bom_components_for_sku,
    create_or_update_sku,
    decimal_to_str,
    ensure_balance,
    get_sku_by_code,
    normalize_sku,
    optional_decimal_to_str,
    pending_commitments_by_sku,
    register_movement,
    save_inventory_count,
    to_decimal,
)


SKU_IMPORT_COLUMNS = ["COD", "DESCRICAO", "UNIDADE", "GRUPO", "CATEGORIA", "SALDO_ATUAL"]
SKU_REQUIRED_COLUMNS = ["COD", "DESCRICAO"]
LOCAL_SKU_SHEET_NAME = "Controle - Cadastros"
LOCAL_SKU_HEADER_ROW = 2
LOCAL_SKU_CODE_ALIASES = ["NOVO_COD", "COD", "SKU"]
LOCAL_SKU_DESCRIPTION_ALIASES = ["DESCRICAO", "DESCRICAO_PRIMARIA"]
LOCAL_SKU_PRIMARY_DESC_ALIASES = ["DESCRICAO_PRIMARIA", "DESCRICAO_PRIMARIA_"]
LOCAL_SKU_SECONDARY_DESC_ALIASES = ["DESCRICAO_SECUNDARIA", "DESCRICAO_SECUNDARIA_"]
LOCAL_SKU_SUFFIX_ALIASES = ["SUFIXO"]
LOCAL_SKU_GROUP_ALIASES = ["GRUPO"]
LOCAL_SKU_CATEGORY_ALIASES = ["CATEGORIA"]
LOCAL_SKU_UNIT_ALIASES = ["UN_MEDI_INTERNA", "UN_MEDI_COMERCIAL", "UNIDADE", "UNIDADE_DE_MEDIDA", "UM"]
LOCAL_SKU_STATUS_ALIASES = ["STATUS", "SITUACAO", "SITUACAO_CADASTRO"]
INACTIVE_STATUS_VALUES = {
    "0",
    "FALSE",
    "NAO",
    "NO",
    "OFF",
    "INATIVO",
    "INATIVA",
    "INATIVADO",
    "INATIVADA",
    "DESATIVADO",
    "DESATIVADA",
    "CANCELADO",
    "CANCELADA",
    "OBSOLETO",
    "OBSOLETA",
    "BLOQUEADO",
    "BLOQUEADA",
}
STOCK_COLUMN_ALIASES = ["SALDO_ATUAL", "ESTOQUE_ATUAL", "ESTOQUE", "SALDO"]
SKU_UNIT_ALIASES = ["UNIDADE", "UNIDADE_DE_MEDIDA", "UNIDADE_MEDIDA", "UM"]
LABEL_IMPORT_COLUMNS = ["COD", "QUANTIDADE"]
CONSUMPTION_IMPORT_COLUMNS = ["COD", "UNIDADE_DE_MEDIDA", "SALDO_CONSUMIDO"]
COMMITMENT_IMPORT_COLUMNS = ["COD", "UNIDADE_DE_MEDIDA", "SALDO_EMPENHADO"]
BOM_IMPORT_COLUMNS = ["ITEM_CODIGO", "COMPONENTE_CODIGO", "DESCRICAO", "UNIDADE", "QUANTIDADE"]
BOM_REQUIRED_COLUMNS = ["ITEM_CODIGO", "COMPONENTE_CODIGO", "DESCRICAO", "UNIDADE", "QUANTIDADE"]
INVENTORY_COUNT_IMPORT_COLUMNS = ["COD", "UNIDADE_DE_MEDIDA", "SALDO_CONTADO"]
INVENTORY_ADD_IMPORT_COLUMNS = ["COD", "UNIDADE_DE_MEDIDA", "SALDO_SOMAR"]
CONSUMPTION_UNIT_ALIASES = ["UNIDADE_DE_MEDIDA", "UNIDADE_MEDIDA", "UNIDADE", "UM"]
CONSUMPTION_QTY_ALIASES = [
    "SALDO_CONSUMIDO",
    "CONSUMO_REAL",
    "QUANTIDADE_CONSUMIDA",
    "QTD_CONSUMIDA",
    "CONSUMIDO",
]
COMMITMENT_QTY_ALIASES = [
    "SALDO_EMPENHADO",
    "QUANTIDADE_EMPENHADA",
    "QTD_EMPENHADA",
    "EMPENHADO",
]
INVENTORY_COUNT_QTY_ALIASES = [
    "SALDO_CONTADO",
    "QUANTIDADE_CONTADA",
    "QTD_CONTADA",
    "CONTAGEM",
    "SALDO_FISICO",
]
INVENTORY_ADD_QTY_ALIASES = [
    "SALDO_SOMAR",
    "QUANTIDADE_SOMAR",
    "QTD_SOMAR",
    "SOMAR_SALDO",
    "ENTRADA_MASSA",
]
MASS_MATERIAL_CODE_ALIASES = ["CODIGO", "COD", "SKU"]
MASS_MATERIAL_OS_ALIASES = ["NUMERO_OS", "N_OS", "OS", "ORDEM_SERVICO", "ORDEM_DE_SERVICO"]
MASS_MATERIAL_ITEM_OS_ALIASES = ["ITEM_OS", "ITEM_DA_OS", "CODIGO_PAI", "ITEM_PAI"]
MASS_MATERIAL_DESC_ALIASES = ["DESCRICAO", "DESCRICAO_PRIMARIA", "DESCRICAO_ITEM"]
MASS_MATERIAL_UNIT_ALIASES = ["UNIDADE", "UNIDADE_DE_MEDIDA", "UNIDADE_MEDIDA", "UN_MEDI_INTERNA", "UM"]
MASS_MATERIAL_SECTOR_ALIASES = ["SETOR", "DEPARTAMENTO", "AREA"]
MASS_MATERIAL_QTY_ALIASES = [
    "QTD",
    "QTD_TOTAL",
    "QUANTIDADE",
    "SALDO_CONSUMIDO",
    "SALDO_EMPENHADO",
    "CONSUMO_REAL",
    "QUANTIDADE_CONSUMIDA",
    "QUANTIDADE_EMPENHADA",
]


def _normalize_header(value):
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_")


def _headers(ws, required_columns, max_scan_rows=25):
    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = {
            _normalize_header(cell.value): idx
            for idx, cell in enumerate(ws[row_number], start=1)
            if cell.value is not None and str(cell.value).strip()
        }
        if "COD" in headers and "SKU" not in headers:
            headers["SKU"] = headers["COD"]
        elif "SKU" in headers and "COD" not in headers:
            headers["COD"] = headers["SKU"]
        if all(column in headers for column in required_columns):
            return headers, row_number
    return {}, 1


def _headers_at_row(ws, row_number):
    headers = {
        _normalize_header(cell.value): idx
        for idx, cell in enumerate(ws[row_number], start=1)
        if cell.value is not None and str(cell.value).strip()
    }
    if "COD" in headers and "SKU" not in headers:
        headers["SKU"] = headers["COD"]
    elif "SKU" in headers and "COD" not in headers:
        headers["COD"] = headers["SKU"]
    return headers


def _cell(ws, row_number, headers, column_name, default=None):
    index = headers.get(column_name)
    if not index:
        return default
    return ws.cell(row_number, index).value


def _first_cell(ws, row_number, headers, column_names, default=None):
    for column_name in column_names:
        value = _cell(ws, row_number, headers, column_name)
        if value is not None and str(value).strip() != "":
            return value
    return default


def _find_header(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return alias
    return None


def _first_header(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return alias
    return None


def _first_existing_cell(ws, row_number, headers, column_names, default=None):
    column_name = _first_header(headers, column_names)
    if not column_name:
        return default
    return _cell(ws, row_number, headers, column_name, default)


def _join_description(*parts):
    return " ".join(str(part).strip() for part in parts if part is not None and str(part).strip())


def _status_to_active(value, default=True):
    if value is None or str(value).strip() == "":
        return default
    return _normalize_header(value) not in INACTIVE_STATUS_VALUES


def _normalize_text(value):
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text.upper()).strip()


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"


def _load_xls_workbook(file_obj):
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Arquivos .xls exigem a dependencia xlrd instalada.") from exc

    if isinstance(file_obj, (str, Path)):
        book = xlrd.open_workbook(str(file_obj))
    else:
        book = xlrd.open_workbook(file_contents=file_obj.read())

    wb = Workbook()
    for sheet_index, sheet in enumerate(book.sheets()):
        ws = wb.active if sheet_index == 0 else wb.create_sheet()
        ws.title = sheet.name[:31] or f"Planilha {sheet_index + 1}"
        for row_index in range(sheet.nrows):
            ws.append([sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)])
    return wb


def _load_workbook_for_read(file_obj):
    filename = str(getattr(file_obj, "filename", "") or file_obj)
    if Path(filename).suffix.lower() == ".xls":
        return _load_xls_workbook(file_obj)
    return load_workbook(file_obj, data_only=True)


def import_skus_from_excel(db, file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, SKU_REQUIRED_COLUMNS)
    missing = [col for col in SKU_REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"created": 0, "updated": 0, "balances_updated": 0, "errors": []}
    rows = []
    seen_skus = set()
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_desc = ws.cell(row_number, headers["DESCRICAO"]).value
        if not raw_sku and not raw_desc:
            continue

        existing_sku = get_sku_by_code(db, raw_sku)
        data = {
            "sku": raw_sku,
            "descricao": raw_desc,
            "active": existing_sku.active if existing_sku else True,
        }
        optional_fields = {
            "unidade": SKU_UNIT_ALIASES,
            "grupo": ["GRUPO"],
            "categoria": ["CATEGORIA"],
        }
        for field_name, aliases in optional_fields.items():
            value = _first_cell(ws, row_number, headers, aliases)
            if value is not None and str(value).strip() != "":
                data[field_name] = value
        if any(column in headers for column in STOCK_COLUMN_ALIASES):
            saldo_atual = _first_cell(ws, row_number, headers, STOCK_COLUMN_ALIASES)
            if saldo_atual is not None and str(saldo_atual).strip() != "":
                data["saldo_atual"] = saldo_atual

        try:
            sku_code = normalize_sku(raw_sku)
            if not sku_code:
                raise ValueError("COD e obrigatorio.")
            if not str(raw_desc or "").strip():
                raise ValueError("Descricao e obrigatoria.")
            if sku_code in seen_skus:
                raise ValueError("COD duplicado na planilha.")
            seen_skus.add(sku_code)
            if "saldo_atual" in data:
                to_decimal(data.get("saldo_atual"))
            rows.append(data)
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhum COD encontrado na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        for data in rows:
            _, created = create_or_update_sku(db, data, commit=False)
            if created:
                result["created"] += 1
            else:
                result["updated"] += 1
            if "saldo_atual" in data:
                result["balances_updated"] += 1
        db.commit()
    except Exception as exc:
        db.rollback()
        result["created"] = 0
        result["updated"] = 0
        result["balances_updated"] = 0
        result["errors"].append(str(exc))
    return result


def import_skus_from_local_master(db, file_obj):
    wb = _load_workbook_for_read(file_obj)
    ws = wb[LOCAL_SKU_SHEET_NAME] if LOCAL_SKU_SHEET_NAME in wb.sheetnames else wb.active
    headers = _headers_at_row(ws, LOCAL_SKU_HEADER_ROW)

    code_header = _first_header(headers, LOCAL_SKU_CODE_ALIASES)
    missing = []
    if not code_header:
        missing.append("NOVO COD")
    if not (
        _first_header(headers, LOCAL_SKU_DESCRIPTION_ALIASES)
        or _first_header(headers, LOCAL_SKU_PRIMARY_DESC_ALIASES)
    ):
        missing.append("DESCRICAO PRIMARIA")
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {
        "created": 0,
        "updated": 0,
        "balances_updated": 0,
        "status_updated": 0,
        "errors": [],
        "rows": 0,
        "duplicates_skipped": 0,
    }
    rows = []
    seen_skus = set()
    for row_number in range(LOCAL_SKU_HEADER_ROW + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers[code_header]).value
        if raw_sku is None or str(raw_sku).strip() == "":
            continue

        primary = _first_cell(ws, row_number, headers, LOCAL_SKU_PRIMARY_DESC_ALIASES)
        descricao = str(primary or "").strip()
        if not descricao:
            descricao = str(_first_cell(ws, row_number, headers, LOCAL_SKU_DESCRIPTION_ALIASES) or "").strip()

        data = {
            "sku": raw_sku,
            "descricao": descricao,
        }

        if _first_header(headers, LOCAL_SKU_UNIT_ALIASES):
            unidade = _first_existing_cell(ws, row_number, headers, LOCAL_SKU_UNIT_ALIASES)
            data["unidade"] = unidade
        if _first_header(headers, LOCAL_SKU_GROUP_ALIASES):
            grupo = _first_existing_cell(ws, row_number, headers, LOCAL_SKU_GROUP_ALIASES)
            data["grupo"] = grupo
        if _first_header(headers, LOCAL_SKU_CATEGORY_ALIASES):
            categoria = _first_existing_cell(ws, row_number, headers, LOCAL_SKU_CATEGORY_ALIASES)
            data["categoria"] = categoria
        status = _first_cell(ws, row_number, headers, LOCAL_SKU_STATUS_ALIASES)
        if status is not None and str(status).strip() != "":
            data["active"] = _status_to_active(status)

        try:
            sku_code = normalize_sku(raw_sku)
            if not sku_code:
                raise ValueError("COD e obrigatorio.")
            if not descricao:
                raise ValueError("Descricao e obrigatoria.")
            if sku_code in seen_skus:
                result["duplicates_skipped"] += 1
                continue
            seen_skus.add(sku_code)
            rows.append(data)
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhum COD encontrado na planilha mestre.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        existing_by_code = {
            normalize_sku(sku.sku): sku
            for sku in db.query(SKU).filter(SKU.sku.in_([normalize_sku(data["sku"]) for data in rows])).all()
        }
        for data in rows:
            sku_code = normalize_sku(data["sku"])
            sku = existing_by_code.get(sku_code)
            created = sku is None
            if created:
                sku = SKU(sku=sku_code, descricao=data["descricao"], active=data.get("active", True))
                db.add(sku)
                existing_by_code[sku_code] = sku
            else:
                sku.descricao = data["descricao"]
                if "active" in data and sku.active != data["active"]:
                    sku.active = data["active"]
                    result["status_updated"] += 1
            if "unidade" in data:
                sku.unidade = str(data.get("unidade") or "").strip() or None
            if "grupo" in data:
                sku.grupo = str(data.get("grupo") or "").strip() or None
            if "categoria" in data:
                sku.categoria = str(data.get("categoria") or "").strip() or None
            db.flush()
            if created:
                ensure_balance(db, sku)
            if created:
                result["created"] += 1
            else:
                result["updated"] += 1
        db.commit()
        result["rows"] = len(rows)
    except Exception as exc:
        db.rollback()
        result["created"] = 0
        result["updated"] = 0
        result["rows"] = 0
        result["errors"].append(str(exc))
    return result


def import_label_jobs_from_excel(db, file_obj, user_id, inventory_session_id=None):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, LABEL_IMPORT_COLUMNS)
    missing = [col for col in LABEL_IMPORT_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"created": 0, "errors": []}
    origem = "INVENTARIO" if inventory_session_id else "EXCEL"
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_qty = ws.cell(row_number, headers["QUANTIDADE"]).value or 1
        if not raw_sku:
            continue
        sku = get_sku_by_code(db, raw_sku)
        try:
            if not sku:
                raise ValueError("COD nao cadastrado")
            if not sku.active:
                raise ValueError("COD inativo")
            qty = int(to_decimal(raw_qty))
            if qty <= 0:
                raise ValueError("Quantidade deve ser maior que zero")
            create_label_job(db, sku, qty, origem, user_id, inventory_session_id)
            result["created"] += 1
        except Exception as exc:
            db.rollback()
            result["errors"].append(f"Linha {row_number}: {raw_sku} - {exc}")
    return result


def import_consumption_from_excel(db, file_obj, user_id, documento="", observacao="", allow_negative=False):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, ["COD"])
    qty_header = _find_header(headers, CONSUMPTION_QTY_ALIASES)
    unit_header = _find_header(headers, CONSUMPTION_UNIT_ALIASES)
    missing = []
    if "COD" not in headers:
        missing.append("COD")
    if not unit_header:
        missing.append("UNIDADE_DE_MEDIDA")
    if not qty_header:
        missing.append("SALDO_CONSUMIDO")
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"processed": 0, "total_consumed": "0", "errors": []}
    rows = []
    total_consumed = to_decimal(0)
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_unit = ws.cell(row_number, headers[unit_header]).value
        raw_consumed = ws.cell(row_number, headers[qty_header]).value
        if not raw_sku and not raw_unit and not raw_consumed:
            continue

        try:
            sku = get_sku_by_code(db, raw_sku, active_only=True)
            if not sku:
                raise ValueError("COD nao cadastrado ou inativo")
            if raw_unit is None or str(raw_unit).strip() == "":
                raise ValueError("Unidade de medida e obrigatoria")
            if sku.unidade and _normalize_text(raw_unit) != _normalize_text(sku.unidade):
                raise ValueError(f"Unidade divergente. Cadastro: {sku.unidade}")
            consumed = to_decimal(raw_consumed)
            if consumed <= 0:
                raise ValueError("Saldo consumido deve ser maior que zero")
            if not allow_negative:
                saldo_atual = to_decimal(sku.balance.saldo_atual if sku.balance else 0)
                if saldo_atual - consumed < 0:
                    raise ValueError("Saldo insuficiente para baixa")
            rows.append((row_number, sku, str(raw_unit).strip(), consumed))
            total_consumed += consumed
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {raw_sku or ''} - {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhum consumo encontrado na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        document = documento or f"BAIXA-EXCEL-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        for row_number, sku, unidade, consumed in rows:
            note = observacao or f"Baixa por consumo real via planilha. Linha {row_number}; unidade {unidade}."
            register_movement(
                db,
                sku,
                "BAIXA",
                consumed,
                user_id,
                documento=document,
                observacao=note,
                allow_negative=allow_negative,
                commit=False,
            )
            result["processed"] += 1
        db.commit()
        result["total_consumed"] = decimal_to_str(total_consumed)
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result["total_consumed"] = "0"
        result["errors"].append(str(exc))
    return result


def import_commitments_from_excel(db, file_obj, user_id, documento="", observacao=""):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, ["COD"])
    qty_header = _find_header(headers, COMMITMENT_QTY_ALIASES)
    unit_header = _find_header(headers, CONSUMPTION_UNIT_ALIASES)
    missing = []
    if "COD" not in headers:
        missing.append("COD")
    if not unit_header:
        missing.append("UNIDADE_DE_MEDIDA")
    if not qty_header:
        missing.append("SALDO_EMPENHADO")
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"processed": 0, "total_committed": "0", "errors": []}
    rows = []
    total_committed = to_decimal(0)
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_unit = ws.cell(row_number, headers[unit_header]).value
        raw_committed = ws.cell(row_number, headers[qty_header]).value
        if not raw_sku and not raw_unit and not raw_committed:
            continue

        try:
            sku = get_sku_by_code(db, raw_sku, active_only=True)
            if not sku:
                raise ValueError("COD nao cadastrado ou inativo")
            if raw_unit is None or str(raw_unit).strip() == "":
                raise ValueError("Unidade de medida e obrigatoria")
            if sku.unidade and _normalize_text(raw_unit) != _normalize_text(sku.unidade):
                raise ValueError(f"Unidade divergente. Cadastro: {sku.unidade}")
            committed = to_decimal(raw_committed)
            if committed <= 0:
                raise ValueError("Saldo empenhado deve ser maior que zero")
            rows.append((row_number, sku, str(raw_unit).strip(), committed))
            total_committed += committed
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {raw_sku or ''} - {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhum empenho encontrado na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        document = documento or f"EMPENHO-EXCEL-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        for row_number, sku, unidade, committed in rows:
            note = observacao or f"Empenho inicial via planilha. Linha {row_number}; unidade {unidade}."
            register_movement(
                db,
                sku,
                "EMPENHO",
                committed,
                user_id,
                documento=document,
                observacao=note,
                commit=False,
            )
            result["processed"] += 1
        db.commit()
        result["total_committed"] = decimal_to_str(total_committed)
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result["total_committed"] = "0"
        result["errors"].append(str(exc))
    return result


def _headers_with_code(ws, max_scan_rows=25):
    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = _headers_at_row(ws, row_number)
        if _first_header(headers, MASS_MATERIAL_CODE_ALIASES):
            return headers, row_number
    return {}, 1


def _requisition_worksheet(wb):
    for sheet_name in ("Requisicao", "Requisição", "Somatorio", "Somatório"):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
    return wb.active


def _has_quantity(value):
    return value is not None and str(value).strip() != ""


def parse_mass_materials_from_excel(file_obj):
    wb = _load_workbook_for_read(file_obj)
    ws = _requisition_worksheet(wb)
    headers, header_row = _headers_with_code(ws)
    code_header = _first_header(headers, MASS_MATERIAL_CODE_ALIASES)
    qty_header = _first_header(headers, MASS_MATERIAL_QTY_ALIASES)
    if not code_header:
        raise ValueError("Coluna CODIGO ausente.")

    rows = []
    missing_quantities = []
    errors = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_code = _cell(ws, row_number, headers, code_header)
        raw_desc = _first_existing_cell(ws, row_number, headers, MASS_MATERIAL_DESC_ALIASES, "")
        raw_unit = _first_existing_cell(ws, row_number, headers, MASS_MATERIAL_UNIT_ALIASES, "")
        raw_sector = _first_existing_cell(ws, row_number, headers, MASS_MATERIAL_SECTOR_ALIASES, "")
        raw_os = _first_existing_cell(ws, row_number, headers, MASS_MATERIAL_OS_ALIASES, "")
        raw_item_os = _first_existing_cell(ws, row_number, headers, MASS_MATERIAL_ITEM_OS_ALIASES, "")
        raw_qty = _cell(ws, row_number, headers, qty_header) if qty_header else None

        if not raw_code and not raw_desc and not raw_unit and not raw_qty:
            continue
        code = normalize_sku(raw_code)
        if not code:
            errors.append(f"Linha {row_number}: CODIGO e obrigatorio.")
            continue

        row = {
            "linha": row_number,
            "numero_os": str(raw_os or "").strip(),
            "item_os": normalize_sku(raw_item_os),
            "codigo": code,
            "descricao": str(raw_desc or "").strip(),
            "unidade": str(raw_unit or "").strip(),
            "setor": str(raw_sector or "").strip(),
            "quantidade": "" if raw_qty is None else str(raw_qty).strip(),
            "missing_quantity": not _has_quantity(raw_qty),
        }
        if not _has_quantity(raw_qty):
            missing_quantities.append(row)
        else:
            try:
                qty = to_decimal(raw_qty)
                if qty <= 0:
                    raise ValueError("Quantidade deve ser maior que zero.")
            except Exception as exc:
                errors.append(f"Linha {row_number}: {code} - {exc}")
        rows.append(row)

    if not rows and not errors:
        errors.append("Nenhum CODIGO encontrado na planilha.")
    return {"rows": rows, "missing_quantities": missing_quantities, "errors": errors}


def mass_material_rows_from_form(form):
    rows = []
    codes = form.getlist("codigo")
    for index, code in enumerate(codes):
        rows.append(
            {
                "linha": form.getlist("linha")[index] if index < len(form.getlist("linha")) else "",
                "numero_os": form.getlist("numero_os")[index] if index < len(form.getlist("numero_os")) else "",
                "item_os": form.getlist("item_os")[index] if index < len(form.getlist("item_os")) else "",
                "codigo": code,
                "descricao": form.getlist("descricao")[index] if index < len(form.getlist("descricao")) else "",
                "unidade": form.getlist("unidade")[index] if index < len(form.getlist("unidade")) else "",
                "setor": form.getlist("setor")[index] if index < len(form.getlist("setor")) else "",
                "quantidade": form.getlist("quantidade")[index] if index < len(form.getlist("quantidade")) else "",
            }
        )
    return rows


def _validate_unit_from_registration(sku, raw_unit):
    return


def is_preparation_sector(value):
    return _normalize_header(value).startswith("PREPARA")


def skip_preparation_rows_for_consumption(preview):
    skipped = 0
    for row in preview.get("rows", []):
        if is_preparation_sector(row.get("setor")):
            skipped += 1
            row["missing_quantity"] = False
    preview["missing_quantities"] = [
        row
        for row in preview.get("rows", [])
        if row.get("missing_quantity") and not is_preparation_sector(row.get("setor"))
    ]
    preview["skipped_preparation"] = skipped
    return preview


def _expand_mass_material_row(db, sku, qty, row, explicit_parent_codes, errors, visited=None):
    visited = visited or set()
    sku_code = normalize_sku(sku.sku)
    if sku_code in visited:
        errors.append(f"Linha {row.get('linha')}: {sku_code} - B.O.M circular detectada.")
        return [], 0

    components = bom_components_for_sku(db, sku)
    if components:
        if sku_code in explicit_parent_codes:
            return [], 1
        expanded = []
        skipped = 0
        for component in components:
            component_sku = component.component_sku
            if not component_sku or not component_sku.active:
                errors.append(f"Linha {row.get('linha')}: componente inativo ou nao cadastrado na B.O.M de {sku_code}.")
                continue
            component_qty = to_decimal(component.quantidade) * qty
            child_row = dict(row)
            child_row["codigo"] = component_sku.sku
            child_row["unidade"] = component_sku.unidade or component.unidade or ""
            child_row["descricao"] = component.descricao or component_sku.descricao
            child_expanded, child_skipped = _expand_mass_material_row(
                db,
                component_sku,
                component_qty,
                child_row,
                explicit_parent_codes,
                errors,
                visited | {sku_code},
            )
            expanded.extend(child_expanded)
            skipped += child_skipped
        return expanded, skipped

    _validate_unit_from_registration(sku, row.get("unidade"))
    final_row = {
        "sku": sku,
        "quantidade": qty,
        "numero_os": str(row.get("numero_os") or "").strip(),
        "linha": row.get("linha"),
        "origem_codigo": row.get("codigo"),
        "unidade": sku.unidade or row.get("unidade") or "",
    }
    return [final_row], 0


def import_mass_material_movements(db, rows, mode, user_id, documento="", observacao="", allow_negative=False):
    movement_type = "BAIXA" if mode == "BAIXA" else "EMPENHO"
    total_key = "total_consumed" if movement_type == "BAIXA" else "total_committed"
    result = {
        "processed": 0,
        total_key: "0",
        "errors": [],
        "expanded_components": 0,
        "skipped_assemblies": 0,
        "skipped_preparation": 0,
        "os_numbers": [],
    }
    explicit_parent_codes = {normalize_sku(row.get("item_os")) for row in rows if normalize_sku(row.get("item_os"))}
    final_rows = []
    total = to_decimal(0)
    os_numbers = set()

    for row in rows:
        code = normalize_sku(row.get("codigo"))
        try:
            if row.get("numero_os"):
                os_numbers.add(str(row.get("numero_os")).strip())
            if movement_type == "BAIXA" and is_preparation_sector(row.get("setor")):
                result["skipped_preparation"] += 1
                continue
            if not code:
                raise ValueError("CODIGO e obrigatorio.")
            qty = to_decimal(row.get("quantidade"))
            if qty <= 0:
                raise ValueError("Quantidade deve ser maior que zero.")
            sku = get_sku_by_code(db, code, active_only=True)
            if not sku:
                raise ValueError("COD nao cadastrado ou inativo.")
            expanded_rows, skipped = _expand_mass_material_row(db, sku, qty, row, explicit_parent_codes, result["errors"])
            result["skipped_assemblies"] += skipped
            final_rows.extend(expanded_rows)
        except Exception as exc:
            result["errors"].append(f"Linha {row.get('linha')}: {code or ''} - {exc}")

    if result["errors"]:
        db.rollback()
        return result
    if not final_rows:
        if movement_type == "BAIXA" and result["skipped_preparation"]:
            result["os_numbers"] = sorted(os_numbers)
            return result
        result["errors"].append("Nenhum item consumivel encontrado para movimentar.")
        db.rollback()
        return result

    aggregated = {}
    for row in final_rows:
        sku = row["sku"]
        os_number = row["numero_os"]
        key = (sku.id, os_number)
        if key not in aggregated:
            aggregated[key] = dict(row)
        else:
            aggregated[key]["quantidade"] += row["quantidade"]

    try:
        default_document = documento or f"{movement_type}-OS-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        for row in aggregated.values():
            sku = row["sku"]
            qty = to_decimal(row["quantidade"])
            os_number = row.get("numero_os")
            document = f"OS {os_number}" if os_number else default_document
            note = observacao or f"{movement_type} em massa por requisicao de materiais."
            if row.get("origem_codigo") and row["origem_codigo"] != sku.sku:
                note = f"{note} Origem/conjunto: {row['origem_codigo']}."
            register_movement(
                db,
                sku,
                movement_type,
                qty,
                user_id,
                documento=document,
                observacao=note,
                allow_negative=allow_negative,
                commit=False,
            )
            result["processed"] += 1
            total += qty
        db.commit()
        result[total_key] = decimal_to_str(total)
        result["expanded_components"] = len(final_rows) - len(rows) + result["skipped_assemblies"]
        result["os_numbers"] = sorted(os_numbers)
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result[total_key] = "0"
        result["errors"].append(str(exc))
    return result


def _active_sku_from_cache(db, raw_code, sku_cache=None):
    if sku_cache is not None:
        return sku_cache.get(normalize_sku(raw_code))
    return get_sku_by_code(db, raw_code, active_only=True)


def _invalid_bom_code(raw_code):
    code = normalize_sku(raw_code)
    return code in {"", "0", "#N/A", "N/A", "NA", "NONE", "NULL"}


def _parse_bom_workbook(db, file_obj, source_label="", seen_pairs=None, sku_cache=None, lenient=False):
    wb = _load_workbook_for_read(file_obj)
    ws = wb.active
    headers, header_row = _headers(ws, BOM_REQUIRED_COLUMNS)
    missing = [col for col in BOM_REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    rows = []
    parent_ids = set()
    errors = []
    seen_pairs = seen_pairs if seen_pairs is not None else {}
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_item = ws.cell(row_number, headers["ITEM_CODIGO"]).value
        raw_component = ws.cell(row_number, headers["COMPONENTE_CODIGO"]).value
        raw_desc = ws.cell(row_number, headers["DESCRICAO"]).value
        raw_unit = ws.cell(row_number, headers["UNIDADE"]).value
        raw_qty = ws.cell(row_number, headers["QUANTIDADE"]).value
        if not raw_item and not raw_component and not raw_desc and not raw_unit and not raw_qty:
            continue

        try:
            if _invalid_bom_code(raw_item) or _invalid_bom_code(raw_component):
                if lenient:
                    continue
                raise ValueError("Item pai e componente sao obrigatorios")
            item_sku = _active_sku_from_cache(db, raw_item, sku_cache)
            if not item_sku:
                raise ValueError("Item pai nao cadastrado ou inativo")
            component_sku = _active_sku_from_cache(db, raw_component, sku_cache)
            if not component_sku:
                raise ValueError("Componente nao cadastrado ou inativo")
            if item_sku.id == component_sku.id:
                raise ValueError("Item pai e componente nao podem ser iguais")
            descricao = str(raw_desc or "").strip()
            if not descricao:
                raise ValueError("Descricao e obrigatoria")
            unidade = str(component_sku.unidade or "").strip() or str(raw_unit or "").strip()
            quantidade = to_decimal(raw_qty)
            if quantidade <= 0:
                if lenient:
                    continue
                raise ValueError("Quantidade deve ser maior que zero")
            pair = (item_sku.id, component_sku.id)
            if pair in seen_pairs:
                seen_pairs[pair].quantidade = to_decimal(seen_pairs[pair].quantidade) + quantidade
                continue
            parent_ids.add(item_sku.id)
            row = BomComponent(
                item_sku_id=item_sku.id,
                component_sku_id=component_sku.id,
                descricao=descricao,
                unidade=unidade,
                quantidade=quantidade,
            )
            rows.append(row)
            seen_pairs[pair] = row
        except Exception as exc:
            prefix = f"{source_label} - " if source_label else ""
            errors.append(f"{prefix}Linha {row_number}: {raw_item or ''} / {raw_component or ''} - {exc}")
    return rows, parent_ids, errors


def _commit_bom_rows(db, rows, parent_ids, result):
    if not rows and not result["errors"]:
        result["errors"].append("Nenhuma estrutura B.O.M encontrada na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        deleted = (
            db.query(BomComponent)
            .filter(BomComponent.item_sku_id.in_(parent_ids))
            .delete(synchronize_session=False)
        )
        for row in rows:
            db.add(row)
        db.commit()
        result["processed"] = len(rows)
        result["items"] = len(parent_ids)
        result["deleted"] = deleted
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result["items"] = 0
        result["deleted"] = 0
        result["errors"].append(str(exc))
    return result


def import_bom_from_excel(db, file_obj):
    result = {"processed": 0, "items": 0, "deleted": 0, "errors": []}
    sku_cache = {normalize_sku(sku.sku): sku for sku in db.query(SKU).filter(SKU.active.is_(True)).all()}
    rows, parent_ids, errors = _parse_bom_workbook(db, file_obj, sku_cache=sku_cache)
    result["errors"].extend(errors)
    return _commit_bom_rows(db, rows, parent_ids, result)


def import_bom_from_files(db, file_paths):
    result = {"processed": 0, "items": 0, "deleted": 0, "files": 0, "errors": [], "warnings": []}
    rows = []
    parent_ids = set()
    seen_pairs = {}
    sku_cache = {normalize_sku(sku.sku): sku for sku in db.query(SKU).filter(SKU.active.is_(True)).all()}

    for file_path in file_paths:
        path = Path(file_path)
        try:
            parsed_rows, parsed_parent_ids, errors = _parse_bom_workbook(
                db,
                path,
                source_label=path.name,
                seen_pairs=seen_pairs,
                sku_cache=sku_cache,
                lenient=True,
            )
            result["files"] += 1
            rows.extend(parsed_rows)
            parent_ids.update(parsed_parent_ids)
            result["warnings"].extend(errors)
        except Exception as exc:
            result["warnings"].append(f"{path.name}: {exc}")

    if not file_paths:
        result["errors"].append("Nenhuma planilha B.O.M encontrada na pasta local.")

    return _commit_bom_rows(db, rows, parent_ids, result)


def import_inventory_counts_from_excel(db, file_obj, session_id, user_id):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, ["COD"])
    qty_header = _find_header(headers, INVENTORY_COUNT_QTY_ALIASES)
    unit_header = _find_header(headers, CONSUMPTION_UNIT_ALIASES)
    missing = []
    if "COD" not in headers:
        missing.append("COD")
    if not qty_header:
        missing.append("SALDO_CONTADO")
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"processed": 0, "errors": []}
    rows = []
    seen_skus = set()
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_unit = ws.cell(row_number, headers[unit_header]).value if unit_header else ""
        raw_counted = ws.cell(row_number, headers[qty_header]).value
        if not raw_sku and not raw_unit and raw_counted in (None, ""):
            continue

        try:
            sku = get_sku_by_code(db, raw_sku, active_only=True)
            if not sku:
                raise ValueError("COD nao cadastrado ou inativo")
            sku_code = normalize_sku(raw_sku)
            if sku_code in seen_skus:
                raise ValueError("COD duplicado na planilha")
            seen_skus.add(sku_code)
            if raw_unit is not None and str(raw_unit).strip():
                if sku.unidade and _normalize_text(raw_unit) != _normalize_text(sku.unidade):
                    raise ValueError(f"Unidade divergente. Cadastro: {sku.unidade}")
            counted = to_decimal(raw_counted)
            if counted < 0:
                raise ValueError("Saldo contado nao pode ser negativo")
            rows.append((row_number, sku, counted))
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {raw_sku or ''} - {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhuma contagem encontrada na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        for _, sku, counted in rows:
            save_inventory_count(db, session_id, sku, counted, user_id, commit=False)
            result["processed"] += 1
        db.commit()
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result["errors"].append(str(exc))
    return result


def import_inventory_balance_additions_from_excel(db, file_obj, session_id, user_id):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers, header_row = _headers(ws, ["COD"])
    qty_header = _find_header(headers, INVENTORY_ADD_QTY_ALIASES)
    unit_header = _find_header(headers, CONSUMPTION_UNIT_ALIASES)
    missing = []
    if "COD" not in headers:
        missing.append("COD")
    if not qty_header:
        missing.append("SALDO_SOMAR")
    if missing:
        raise ValueError(f"Colunas ausentes: {', '.join(missing)}")

    result = {"processed": 0, "total_added": "0", "errors": []}
    rows = []
    seen_skus = set()
    total_added = to_decimal(0)
    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_sku = ws.cell(row_number, headers["COD"]).value
        raw_unit = ws.cell(row_number, headers[unit_header]).value if unit_header else ""
        raw_addition = ws.cell(row_number, headers[qty_header]).value
        if not raw_sku and not raw_unit and raw_addition in (None, ""):
            continue

        try:
            sku = get_sku_by_code(db, raw_sku, active_only=True)
            if not sku:
                raise ValueError("COD nao cadastrado ou inativo")
            sku_code = normalize_sku(raw_sku)
            if sku_code in seen_skus:
                raise ValueError("COD duplicado na planilha")
            seen_skus.add(sku_code)
            if raw_unit is not None and str(raw_unit).strip():
                if sku.unidade and _normalize_text(raw_unit) != _normalize_text(sku.unidade):
                    raise ValueError(f"Unidade divergente. Cadastro: {sku.unidade}")
            addition = to_decimal(raw_addition)
            if addition <= 0:
                raise ValueError("Saldo a somar deve ser maior que zero")
            saldo_atual = to_decimal(sku.balance.saldo_atual if sku.balance else 0)
            counted = saldo_atual + addition
            rows.append((row_number, sku, counted, addition))
            total_added += addition
        except Exception as exc:
            result["errors"].append(f"Linha {row_number}: {raw_sku or ''} - {exc}")

    if not rows and not result["errors"]:
        result["errors"].append("Nenhum saldo a somar encontrado na planilha.")
    if result["errors"]:
        db.rollback()
        return result

    try:
        for _, sku, counted, _ in rows:
            save_inventory_count(db, session_id, sku, counted, user_id, commit=False)
            result["processed"] += 1
        db.commit()
        result["total_added"] = decimal_to_str(total_added)
    except Exception as exc:
        db.rollback()
        result["processed"] = 0
        result["total_added"] = "0"
        result["errors"].append(str(exc))
    return result


def create_template_files(base_dir):
    template_path = Path(base_dir) / "template_importacao_skus.xlsx"
    sample_path = Path(base_dir) / "dados_exemplo.xlsx"
    label_template_path = Path(base_dir) / "template_etiquetas_lote.xlsx"
    consumption_template_path = Path(base_dir) / "template_baixa_consumo.xlsx"
    commitment_template_path = Path(base_dir) / "template_empenhos.xlsx"
    bom_template_path = Path(base_dir) / "template_bom.xlsx"
    inventory_count_template_path = Path(base_dir) / "template_contagem_inventario.xlsx"
    inventory_add_template_path = Path(base_dir) / "template_somar_saldo_inventario.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Codigos"
    ws.append(SKU_IMPORT_COLUMNS)
    _style_header(ws)
    for width, column in zip([18, 48, 16, 24, 24, 18], "ABCDEF"):
        ws.column_dimensions[column].width = width
    wb.save(template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Codigos"
    ws.append(SKU_IMPORT_COLUMNS)
    examples = [
        ["PAR-0001", "Parafuso sextavado M8 x 30 zincado", "UN", "Ferragens", "Fixadores", 125],
        ["CAB-0012", "Cabo eletrico flexivel 2,5 mm preto", "M", "Componentes", "Eletrica", ""],
        ["CON-0100", "Conector pneumatico reto 1/4", "UN", "Componentes", "Pneumatica", 12],
    ]
    for row in examples:
        ws.append(row)
    _style_header(ws)
    for width, column in zip([18, 48, 16, 24, 24, 18], "ABCDEF"):
        ws.column_dimensions[column].width = width
    wb.save(sample_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Etiquetas"
    ws.append(LABEL_IMPORT_COLUMNS)
    ws.append(["PAR-0001", 2])
    ws.append(["CAB-0012", 1])
    _style_header(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    wb.save(label_template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Baixa"
    ws.append(CONSUMPTION_IMPORT_COLUMNS)
    ws.append(["PAR-0001", "UN", 3])
    ws.append(["CAB-0012", "M", 12.5])
    _style_header(ws)
    for width, column in zip([22, 22, 20], "ABC"):
        ws.column_dimensions[column].width = width
    wb.save(consumption_template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Empenhos"
    ws.append(COMMITMENT_IMPORT_COLUMNS)
    ws.append(["PAR-0001", "UN", 5])
    ws.append(["CAB-0012", "M", 18])
    _style_header(ws)
    for width, column in zip([22, 22, 20], "ABC"):
        ws.column_dimensions[column].width = width
    wb.save(commitment_template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(BOM_IMPORT_COLUMNS)
    ws.append(["PROD-0001", "PAR-0001", "Parafuso sextavado M8 x 30 zincado", "UN", 4])
    ws.append(["PROD-0001", "CAB-0012", "Cabo eletrico flexivel 2,5 mm preto", "M", 2.5])
    _style_header(ws)
    for width, column in zip([22, 24, 48, 14, 18], "ABCDE"):
        ws.column_dimensions[column].width = width
    wb.save(bom_template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contagem"
    ws.append(INVENTORY_COUNT_IMPORT_COLUMNS)
    ws.append(["PAR-0001", "UN", 124])
    ws.append(["CAB-0012", "M", 78.5])
    _style_header(ws)
    for width, column in zip([22, 22, 18], "ABC"):
        ws.column_dimensions[column].width = width
    wb.save(inventory_count_template_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Somar saldo"
    ws.append(INVENTORY_ADD_IMPORT_COLUMNS)
    ws.append(["PAR-0001", "UN", 10])
    ws.append(["CAB-0012", "M", 4.5])
    _style_header(ws)
    for width, column in zip([22, 22, 18], "ABC"):
        ws.column_dimensions[column].width = width
    wb.save(inventory_add_template_path)

    return (
        template_path,
        sample_path,
        label_template_path,
        consumption_template_path,
        commitment_template_path,
        bom_template_path,
        inventory_count_template_path,
        inventory_add_template_path,
    )


def _metadata(ws, title, user, filters=None):
    ws.append([title])
    ws.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    ws.append(["Usuario", user.username if user else ""])
    ws.append(["Filtros", filters or "Sem filtros"])
    ws.append([])


def _autosize(ws):
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 60)


def _save_report(wb, prefix):
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORTS_DIR / f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path


def export_stock_report(db, user, filters):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    _metadata(ws, "Relatorio de estoque atual", user, filters)
    ws.append([
        "COD",
        "Descricao",
        "Unidade",
        "Grupo",
        "Categoria",
        "Localizacao",
        "Saldo atual",
        "Empenhado",
        "Saldo disponivel",
        "Estoque minimo",
        "Ativo",
        "Status",
    ])

    query = db.query(SKU).outerjoin(StockBalance)
    sku_filter = filters.get("sku")
    desc_filter = filters.get("descricao")
    group = filters.get("grupo")
    category = filters.get("categoria")
    location = filters.get("localizacao")
    active = filters.get("active")
    low = filters.get("saldo_baixo")

    if sku_filter:
        query = query.filter(SKU.sku.ilike(f"%{sku_filter}%"))
    if desc_filter:
        query = query.filter(SKU.descricao.ilike(f"%{desc_filter}%"))
    if group:
        query = query.filter(SKU.grupo.ilike(f"%{group}%"))
    if category:
        query = query.filter(SKU.categoria.ilike(f"%{category}%"))
    if location:
        query = query.filter(SKU.localizacao.ilike(f"%{location}%"))
    if active == "1":
        query = query.filter(SKU.active.is_(True))
    elif active == "0":
        query = query.filter(SKU.active.is_(False))
    if low == "1":
        query = query.filter(SKU.estoque_minimo.isnot(None))
        query = query.filter(or_(StockBalance.saldo_atual <= SKU.estoque_minimo, StockBalance.saldo_atual.is_(None)))

    rows = query.order_by(SKU.sku).all()
    pending_by_sku = pending_commitments_by_sku(db, [sku.id for sku in rows])
    for sku in rows:
        saldo = sku.balance.saldo_atual if sku.balance else 0
        empenhado = pending_by_sku.get(sku.id, to_decimal(0))
        disponivel = to_decimal(saldo) - empenhado
        minimo = sku.estoque_minimo
        if to_decimal(saldo) <= 0:
            status = "ZERADO"
        elif minimo is not None and to_decimal(saldo) <= to_decimal(minimo):
            status = "BAIXO"
        else:
            status = "OK"
        ws.append([
            sku.sku,
            sku.descricao,
            sku.unidade,
            sku.grupo,
            sku.categoria,
            sku.localizacao,
            decimal_to_str(saldo),
            decimal_to_str(empenhado),
            decimal_to_str(disponivel),
            optional_decimal_to_str(sku.estoque_minimo),
            "Sim" if sku.active else "Nao",
            status,
        ])
    _autosize(ws)
    return _save_report(wb, "relatorio_estoque")


def export_movements_report(db, user, tipo=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentacoes"
    if isinstance(tipo, (list, tuple, set)):
        tipo_label = ", ".join(tipo)
        title = f"Relatorio de {tipo_label.lower()}"
        filter_label = f"tipo={tipo_label}"
    else:
        title = "Relatorio completo de movimentacoes" if not tipo else f"Relatorio de {tipo.lower()}"
        filter_label = f"tipo={tipo}" if tipo else "Sem filtros"
    _metadata(ws, title, user, filter_label)
    ws.append([
        "ID",
        "Data/Hora",
        "Usuario",
        "COD",
        "Descricao",
        "Tipo",
        "Quantidade",
        "Saldo anterior",
        "Saldo posterior",
        "Empenho origem",
        "Documento",
        "Observacao",
    ])
    query = db.query(Movement).join(SKU)
    if isinstance(tipo, (list, tuple, set)):
        query = query.filter(Movement.tipo.in_(tipo))
    elif tipo:
        query = query.filter(Movement.tipo == tipo)
    for mv in query.order_by(Movement.created_at.desc()).all():
        tipo_display = "EMPENHO" if mv.tipo == "SAIDA" else mv.tipo
        ws.append([
            mv.id,
            mv.created_at.strftime("%d/%m/%Y %H:%M:%S"),
            mv.usuario.username,
            mv.sku.sku,
            mv.sku.descricao,
            tipo_display,
            decimal_to_str(mv.quantidade),
            decimal_to_str(mv.saldo_anterior),
            decimal_to_str(mv.saldo_posterior),
            mv.related_movement_id or "",
            mv.documento,
            mv.observacao,
        ])
    _autosize(ws)
    if isinstance(tipo, (list, tuple, set)):
        prefix = "relatorio_empenhos"
    else:
        prefix = f"relatorio_{tipo.lower() if tipo else 'movimentacoes'}"
    return _save_report(wb, prefix)


def export_inventory_report(db, user, session_id=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    _metadata(ws, "Relatorio de inventario", user, f"sessao={session_id}" if session_id else "Todas as sessoes")
    ws.append([
        "Sessao",
        "Status",
        "COD",
        "Descricao",
        "Saldo sistema",
        "Quantidade contada",
        "Diferenca",
        "Contado por",
        "Contado em",
    ])
    query = db.query(InventoryCount).join(InventorySession)
    if session_id:
        query = query.filter(InventoryCount.session_id == session_id)
    for count in query.order_by(InventoryCount.counted_at.desc()).all():
        ws.append([
            count.session_id,
            count.session.status,
            count.sku.sku,
            count.sku.descricao,
            decimal_to_str(count.saldo_sistema),
            decimal_to_str(count.quantidade_contada),
            decimal_to_str(count.diferenca),
            count.user.username,
            count.counted_at.strftime("%d/%m/%Y %H:%M:%S"),
        ])
    _autosize(ws)
    return _save_report(wb, "relatorio_inventario")


def export_inventory_preview(db, user, session):
    wb = Workbook()
    ws = wb.active
    ws.title = "Previa"
    _metadata(ws, "Previa do inventario", user, f"sessao={session.id if session else ''}")
    ws.append(["COD", "Descricao", "Saldo sistema", "Contagem", "Diferenca", "Status"])

    counts_by_sku = {}
    if session:
        counts_by_sku = {
            count.sku_id: count
            for count in db.query(InventoryCount).filter_by(session_id=session.id).all()
        }
    for sku in db.query(SKU).filter(SKU.active.is_(True)).order_by(SKU.sku).all():
        count = counts_by_sku.get(sku.id)
        saldo = sku.balance.saldo_atual if sku.balance else 0
        ws.append([
            sku.sku,
            sku.descricao,
            decimal_to_str(saldo),
            decimal_to_str(count.quantidade_contada) if count else "",
            decimal_to_str(count.diferenca) if count else "",
            "CONTADO" if count else "PENDENTE",
        ])
    _autosize(ws)
    return _save_report(wb, "previa_inventario")


def label_queue_summary(db, session_id=None):
    query = db.query(LabelPrintJob)
    if session_id:
        query = query.filter(LabelPrintJob.inventory_session_id == session_id)
    rows = query.all()
    return {
        "total": len(rows),
        "pendente": sum(1 for row in rows if row.status == "PENDENTE"),
        "impresso": sum(1 for row in rows if row.status == "IMPRESSO"),
        "erro": sum(1 for row in rows if row.status == "ERRO"),
    }

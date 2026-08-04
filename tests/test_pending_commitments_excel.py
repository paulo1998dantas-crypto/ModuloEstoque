import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker


APP_DIR = Path(__file__).resolve().parents[1] / "estoque_app"
sys.path.insert(0, str(APP_DIR))

from database import Base  # noqa: E402
from models import Movement, SKU, User  # noqa: E402
from services.estoque_service import (  # noqa: E402
    pending_commitment_for_movement,
    register_consumption_from_commitment,
    register_movement,
)
from services.excel_service import (  # noqa: E402
    export_pending_commitments_report,
    import_pending_commitment_consumptions,
    parse_commitment_corrections_from_excel,
    parse_mass_materials_from_excel,
    parse_pending_commitment_consumptions_from_excel,
)


class PendingCommitmentsExcelTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", future=True)
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, future=True)
        self.db = self.Session()
        self.user = User(username="tester", password_hash="hash", role="ADM", active=True)
        self.sku = SKU(sku="MAT-001", descricao="Material de teste", unidade="UN", active=True)
        self.db.add_all([self.user, self.sku])
        self.db.commit()
        register_movement(self.db, self.sku, "ENTRADA", 30, self.user.id)
        self.commitment = register_movement(self.db, self.sku, "EMPENHO", 10, self.user.id)
        register_consumption_from_commitment(
            self.db,
            self.commitment,
            4,
            self.user.id,
        )
        completed = register_movement(self.db, self.sku, "EMPENHO", 2, self.user.id)
        register_consumption_from_commitment(self.db, completed, 2, self.user.id)

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    @staticmethod
    def _header_map(ws):
        for row_number in range(1, 26):
            values = [cell.value for cell in ws[row_number]]
            if "ID_EMPENHO" in values:
                return row_number, {value: index for index, value in enumerate(values, start=1) if value}
        raise AssertionError("Cabecalho ID_EMPENHO nao encontrado")

    def _export(self, directory):
        with patch("services.excel_service.EXPORTS_DIR", Path(directory)):
            return export_pending_commitments_report(self.db, self.user)

    def test_export_contains_only_pending_commitments_and_blank_input_column(self):
        movement_snapshot = [
            (row.id, row.tipo, row.quantidade, row.related_movement_id)
            for row in self.db.query(Movement).order_by(Movement.id).all()
        ]
        balance_before = self.sku.balance.saldo_atual
        with tempfile.TemporaryDirectory() as directory:
            path = self._export(directory)
            workbook = load_workbook(path, data_only=True)
            self.assertIn("Necessidades O.S.", workbook.sheetnames)
            needs_headers = next(
                {
                    cell.value
                    for cell in workbook["Necessidades O.S."][row_number]
                    if cell.value
                }
                for row_number in range(1, 26)
                if "FALTA_EXPEDIR"
                in {
                    cell.value
                    for cell in workbook["Necessidades O.S."][row_number]
                    if cell.value
                }
            )
            self.assertIn("SALDO_EM_FLUXO_NAO_APROPRIADO", needs_headers)
            self.assertIn("EMPENHOS_EM_FLUXO", needs_headers)
            ws = workbook["Empenhos pendentes"]
            header_row, headers = self._header_map(ws)
            rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))

        self.assertEqual(1, len(rows))
        self.assertEqual(self.commitment.id, rows[0][headers["ID_EMPENHO"] - 1])
        self.assertEqual(6, rows[0][headers["SALDO_PENDENTE"] - 1])
        self.assertIsNone(rows[0][headers["EMPENHO"] - 1])
        self.assertEqual(
            movement_snapshot,
            [
                (row.id, row.tipo, row.quantidade, row.related_movement_id)
                for row in self.db.query(Movement).order_by(Movement.id).all()
            ],
        )
        self.assertEqual(balance_before, self.sku.balance.saldo_atual)

    def test_export_formats_shared_commitment_candidates_for_active_work_order(self):
        work_order_id = uuid4().hex
        vehicle_id = uuid4().hex
        entry_id = uuid4().hex
        composition = json.dumps(
            [
                {
                    "codigo": self.sku.sku,
                    "descricao": self.sku.descricao,
                    "unidade": self.sku.unidade,
                    "qtd": 10,
                    "level": 0,
                }
            ]
        )
        with self.engine.begin() as connection:
            connection.execute(
                text(
                    "create table erp_vehicles "
                    "(id text primary key,chassi text not null,marca text,modelo text,versao text)"
                )
            )
            connection.execute(
                text(
                    "create table erp_vehicle_entries "
                    "(id text primary key,vehicle_id text not null,item_number integer not null)"
                )
            )
            connection.execute(
                text(
                    "create table erp_work_orders "
                    "(id text primary key,vehicle_entry_id text not null,numero_os text not null,"
                    "cliente_nome text,status text not null,technical_status text default 'ABERTA')"
                )
            )
            connection.execute(
                text(
                    "create table suprimentos_documentos "
                    "(id integer primary key,tipo text,erp_work_order_id text,numero text,status text,"
                    "composicao text,updated_at text)"
                )
            )
            connection.execute(
                text("insert into erp_vehicles values (:id,:chassi,'JI','Van','V1')"),
                {"id": vehicle_id, "chassi": "9BRTESTE000001234"},
            )
            connection.execute(
                text("insert into erp_vehicle_entries values (:id,:vehicle_id,3100)"),
                {"id": entry_id, "vehicle_id": vehicle_id},
            )
            connection.execute(
                text(
                    "insert into erp_work_orders values "
                    "(:id,:entry_id,'3100','CLIENTE','ATIVA','ABERTA')"
                ),
                {"id": work_order_id, "entry_id": entry_id},
            )
            connection.execute(
                text(
                    "insert into suprimentos_documentos values "
                    "(1,'os',:work_id,'3100','emitido',:composition,'2026-08-03')"
                ),
                {"work_id": work_order_id, "composition": composition},
            )

        with tempfile.TemporaryDirectory() as directory:
            path = self._export(directory)
            workbook = load_workbook(path, data_only=True)
            ws = workbook["Necessidades O.S."]
            header_row = next(
                row_number
                for row_number in range(1, 26)
                if "EMPENHOS_EM_FLUXO"
                in {cell.value for cell in ws[row_number] if cell.value}
            )
            headers = {
                cell.value: index
                for index, cell in enumerate(ws[header_row], start=1)
                if cell.value
            }
            shared_commitments = ws.cell(
                header_row + 1, headers["EMPENHOS_EM_FLUXO"]
            ).value

        self.assertIn(f"#{self.commitment.id}", shared_commitments)
        self.assertIn(self.sku.sku, shared_commitments)

    def test_exported_workbook_can_be_edited_for_audited_correction_preview(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._export(directory)
            wb = load_workbook(path)
            ws = wb["Empenhos pendentes"]
            header_row, headers = self._header_map(ws)
            ws.cell(header_row + 1, headers["QUANTIDADE_EMPENHADA"]).value = 8
            ws.cell(header_row + 1, headers["DOCUMENTO_EMPENHO"]).value = "OS 3100"
            ws.cell(header_row + 1, headers["OBSERVACAO_EMPENHO"]).value = "Reconciliado"
            ws.cell(header_row + 1, headers["ACAO_CORRECAO"]).value = "CORRIGIR"
            ws.cell(header_row + 1, headers["MOTIVO_CORRECAO"]).value = "Marco zero"
            wb.save(path)

            parsed = parse_commitment_corrections_from_excel(path)

        self.assertEqual([], parsed["errors"])
        self.assertEqual(1, len(parsed["rows"]))
        self.assertEqual(self.commitment.id, parsed["rows"][0]["movement_id"])
        self.assertEqual(8, parsed["rows"][0]["quantidade_empenhada"])
        self.assertEqual("OS 3100", parsed["rows"][0]["documento_empenho"])
        self.assertEqual("CORRIGIR", parsed["rows"][0]["acao_correcao"])

    def test_same_exported_file_imports_linked_partial_consumption(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._export(directory)
            wb = load_workbook(path)
            ws = wb["Empenhos pendentes"]
            header_row, headers = self._header_map(ws)
            ws.cell(header_row + 1, headers["EMPENHO"]).value = 3
            wb.save(path)

            preview = parse_pending_commitment_consumptions_from_excel(path)
            result = import_pending_commitment_consumptions(
                self.db,
                preview["rows"],
                self.user.id,
            )

        self.assertEqual([], preview["errors"])
        self.assertEqual(1, result["processed"])
        self.assertEqual("3", result["total_consumed"])
        self.assertEqual(3, pending_commitment_for_movement(self.db, self.commitment))
        linked = self.db.query(Movement).filter_by(
            tipo="BAIXA",
            related_movement_id=self.commitment.id,
        ).all()
        self.assertEqual(2, len(linked))

    def test_import_is_atomic_when_quantity_exceeds_pending_balance(self):
        rows = [
            {
                "linha": 7,
                "movement_id": self.commitment.id,
                "codigo": self.sku.sku,
                "quantidade": 7,
                "documento": "",
                "observacao": "",
            }
        ]
        before = self.db.query(Movement).count()
        result = import_pending_commitment_consumptions(self.db, rows, self.user.id)

        self.assertEqual(0, result["processed"])
        self.assertTrue(result["errors"])
        self.assertIn("excede o saldo pendente 6", result["errors"][0])
        self.assertEqual(before, self.db.query(Movement).count())

    def test_blank_empenho_column_requests_at_least_one_quantity(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._export(directory)
            preview = parse_pending_commitment_consumptions_from_excel(path)

        self.assertEqual([], preview["rows"])
        self.assertEqual(
            ["Preencha a coluna EMPENHO em pelo menos uma linha para realizar a baixa."],
            preview["errors"],
        )

    def test_legacy_consumption_layout_is_left_for_existing_importer(self):
        workbook = Workbook()
        workbook.active.append(["COD", "UNIDADE_DE_MEDIDA", "SALDO_CONSUMIDO"])
        workbook.active.append([self.sku.sku, "UN", 1])
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "baixa_antiga.xlsx"
            workbook.save(path)
            preview = parse_pending_commitment_consumptions_from_excel(path)
            legacy_preview = parse_mass_materials_from_excel(path)

        self.assertIsNone(preview)
        self.assertEqual([], legacy_preview["errors"])
        self.assertEqual(self.sku.sku, legacy_preview["rows"][0]["codigo"])
        self.assertEqual("1", legacy_preview["rows"][0]["quantidade"])


if __name__ == "__main__":
    unittest.main()
